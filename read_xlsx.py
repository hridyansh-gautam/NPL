import io
import openpyxl

from pypdf import PdfReader

attached = PdfReader("ex_xlsx.pdf").attachments  

if len(attached) != 1:
    raise RuntimeError("Expect a single attachment")

for file_name, content_list in attached.items():

    # Elements in content_list are Python byte-literals
    # Assuming byte_string contains the byte data of the XLSX file
    xlsx_data = io.BytesIO( content_list[0] )
    
    # Load the Excel workbook from the byte data
    workbook = openpyxl.load_workbook(xlsx_data)
    
    # Access the sheets in the workbook
    sheets = workbook.sheetnames

    # Iterate over each sheet and extract data
    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        print(f"Sheet: {sheet_name}")
        for row in sheet.iter_rows(values_only=True):
            print(row)
        
